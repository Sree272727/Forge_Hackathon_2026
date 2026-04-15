"""Generate the Dealership financial model test workbook.

Designed to trigger every Rosetta capability:
 - 6 sheets with cross-sheet references
 - 8+ named ranges
 - intentional circular reference (Service Absorption <-> P&L Overhead)
 - 3 stale assumptions (dates > 12 months old)
 - 2 hidden rows with deprecated assumptions
 - hardcoded anomaly in Used Vehicle row 23
 - 200+ formula cells
"""
from __future__ import annotations

import random
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


random.seed(42)

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _header(ws, row, col, text, bold=True):
    ws.cell(row=row, column=col, value=text)
    cell = ws.cell(row=row, column=col)
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal="center")


def build_assumptions(wb: Workbook) -> None:
    ws = wb.create_sheet("Assumptions")
    headers = ["Assumption", "Value", "Effective Date", "Notes"]
    for i, h in enumerate(headers, 1):
        _header(ws, 1, i, h)
    rows = [
        ("FloorPlanRate", 0.058, date(2025, 6, 1), "Floor plan interest annual rate"),
        ("IncentiveRateToyota", 0.021, date(2025, 8, 15), "OEM incentive for Toyota"),
        ("IncentiveRateHonda", 0.018, date(2025, 8, 15), "OEM incentive for Honda"),
        ("IncentiveRateFord", 0.025, date(2025, 8, 15), "OEM incentive for Ford"),
        ("FI_PVR_Target", 1600, date(2025, 1, 10), "Target finance & insurance PVR"),
        ("ServiceAbsorptionTarget", 0.80, date(2023, 2, 10), "Target absorption — STALE"),
        ("TaxRate", 0.21, date(2023, 4, 1), "Federal tax rate — STALE"),
        ("OverheadAllocationMethod", "revenue-weighted", date(2024, 12, 31), "How overhead is allocated"),
        ("OwnerCompensationAddback", 8000, date(2024, 11, 15), "EBITDA addback"),
        ("OneTimeLegalCosts", 5000, date(2024, 11, 15), "EBITDA addback"),
        ("DEPRECATED_FloorPlanRateOld", 0.065, date(2022, 6, 1), "Old rate — DO NOT USE"),
        ("DEPRECATED_HoldbackRate", 0.03, date(2022, 1, 1), "Old holdback — hidden"),
        ("OverheadBase", 120000, date(2024, 1, 1), "Base overhead for absorption calc"),
    ]
    for r, (name, val, dt, note) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=3, value=dt)
        ws.cell(row=r, column=4, value=note)

    # Hide deprecated rows (rows 12 and 13)
    ws.row_dimensions[12].hidden = True
    ws.row_dimensions[13].hidden = True

    # Define workbook-scoped named ranges
    defs = [
        ("FloorPlanRate", "Assumptions!$B$2"),
        ("IncentiveRateToyota", "Assumptions!$B$3"),
        ("IncentiveRateHonda", "Assumptions!$B$4"),
        ("IncentiveRateFord", "Assumptions!$B$5"),
        ("FI_PVR_Target", "Assumptions!$B$6"),
        ("ServiceAbsorptionTarget", "Assumptions!$B$7"),
        ("TaxRate", "Assumptions!$B$8"),
        ("OwnerCompensationAddback", "Assumptions!$B$10"),
        ("OneTimeLegalCosts", "Assumptions!$B$11"),
        ("OverheadBase", "Assumptions!$B$14"),
    ]
    for name, target in defs:
        wb.defined_names[name] = DefinedName(name=name, attr_text=target)


def _write_monthly_header(ws, row=1):
    _header(ws, row, 1, "Line Item")
    for i, m in enumerate(MONTHS, start=2):
        _header(ws, row, i, m)
    _header(ws, row, 14, "YTD")
    _header(ws, row, 15, "Budget")


def build_new_vehicle(wb: Workbook) -> None:
    ws = wb.create_sheet("New Vehicle")
    # Deals table
    headers = ["Stock#", "VIN", "Sale Date", "Sale Price", "Invoice Cost", "Holdback",
               "Incentive Allocation", "Front Gross", "F&I Revenue", "Month", "OEM"]
    for i, h in enumerate(headers, 1):
        _header(ws, 1, i, h)
    oems = ["Toyota", "Honda", "Ford"]
    num_deals = 40
    for r in range(2, 2 + num_deals):
        stock = 1000 + r
        vin = f"VIN{stock:08d}"
        month_idx = (r - 2) % 12
        sale_date = date(2026, month_idx + 1, 5 + (r % 20))
        sale_price = 28000 + random.randint(-4000, 8000)
        invoice = sale_price - random.randint(1500, 3500)
        holdback = round(invoice * 0.02, 2)
        oem = oems[r % 3]
        ws.cell(row=r, column=1, value=stock)
        ws.cell(row=r, column=2, value=vin)
        ws.cell(row=r, column=3, value=sale_date)
        ws.cell(row=r, column=4, value=sale_price)
        ws.cell(row=r, column=5, value=invoice)
        ws.cell(row=r, column=6, value=holdback)
        # Incentive: =Sale Price * IncentiveRate<OEM>
        ws.cell(row=r, column=7, value=f"=D{r}*IncentiveRate{oem}")
        # Front Gross = Sale Price - Invoice + Holdback + Incentive
        ws.cell(row=r, column=8, value=f"=D{r}-E{r}+F{r}+G{r}")
        ws.cell(row=r, column=9, value=round(1400 + random.randint(-300, 800), 2))
        ws.cell(row=r, column=10, value=MONTHS[month_idx])
        ws.cell(row=r, column=11, value=oem)

    # Summary block
    summary_row = 45
    _header(ws, summary_row, 1, "SUMMARY")
    ws.cell(row=summary_row + 1, column=1, value="Unit Count")
    ws.cell(row=summary_row + 1, column=2, value=f"=COUNTA(A2:A{num_deals+1})")
    ws.cell(row=summary_row + 2, column=1, value="Total Front Gross")
    ws.cell(row=summary_row + 2, column=2, value=f"=SUM(H2:H{num_deals+1})")
    ws.cell(row=summary_row + 3, column=1, value="Avg Front Gross")
    ws.cell(row=summary_row + 3, column=2, value=f"=AVERAGE(H2:H{num_deals+1})")
    ws.cell(row=summary_row + 4, column=1, value="Total F&I Revenue")
    ws.cell(row=summary_row + 4, column=2, value=f"=SUM(I2:I{num_deals+1})")
    # Monthly front gross — referenced by P&L Summary
    for i, m in enumerate(MONTHS):
        ws.cell(row=summary_row + 6 + i, column=1, value=f"Front Gross {m}")
        ws.cell(row=summary_row + 6 + i, column=2, value=f"=SUMIFS(H2:H{num_deals+1},J2:J{num_deals+1},\"{m}\")")
    # Monthly total new vehicle gross (= front gross, for P&L) at row summary_row + 20
    ws.cell(row=summary_row + 20, column=1, value="New Vehicle Total Gross")
    for i in range(12):
        col = get_column_letter(2 + i)
        ws.cell(row=summary_row + 20, column=2 + i,
                value=f"=SUMIFS(H2:H{num_deals+1},J2:J{num_deals+1},\"{MONTHS[i]}\")")


def build_used_vehicle(wb: Workbook) -> None:
    ws = wb.create_sheet("Used Vehicle")
    headers = ["Stock#", "VIN", "Sale Date", "Sale Price", "Acquisition Cost",
               "Recon Cost", "Days on Lot", "Floor Plan Interest",
               "Adjusted Gross", "Month", "Source"]
    for i, h in enumerate(headers, 1):
        _header(ws, 1, i, h)
    num_deals = 30
    for r in range(2, 2 + num_deals):
        stock = 2000 + r
        sale_date = date(2026, ((r - 2) % 12) + 1, 10 + (r % 15))
        sale_price = 18000 + random.randint(-3000, 7000)
        acq_cost = sale_price - random.randint(2000, 4500)
        recon = random.randint(300, 1500)
        dol = random.randint(15, 90)
        ws.cell(row=r, column=1, value=stock)
        ws.cell(row=r, column=2, value=f"VIN_U{stock:07d}")
        ws.cell(row=r, column=3, value=sale_date)
        ws.cell(row=r, column=4, value=sale_price)
        ws.cell(row=r, column=5, value=acq_cost)
        ws.cell(row=r, column=6, value=recon)
        ws.cell(row=r, column=7, value=dol)
        # Floor Plan Interest = DaysOnLot * AcquisitionCost * FloorPlanRate / 365
        ws.cell(row=r, column=8, value=f"=G{r}*E{r}*FloorPlanRate/365")
        # Adjusted Gross = Sale Price - Acq - Recon - Floor Plan Interest
        if r == 23:
            # INTENTIONAL ANOMALY: hardcoded value instead of formula
            ws.cell(row=r, column=9, value=3200)
        else:
            ws.cell(row=r, column=9, value=f"=D{r}-E{r}-F{r}-H{r}")
        ws.cell(row=r, column=10, value=MONTHS[(r - 2) % 12])
        ws.cell(row=r, column=11, value=random.choice(["Trade-in", "Auction", "Direct"]))

    summary = 40
    _header(ws, summary, 1, "SUMMARY")
    ws.cell(row=summary + 1, column=1, value="Total Adjusted Gross")
    ws.cell(row=summary + 1, column=2, value=f"=SUM(I2:I{num_deals+1})")
    ws.cell(row=summary + 2, column=1, value="Weighted Avg DOL")
    ws.cell(row=summary + 2, column=2, value=f"=SUMPRODUCT(G2:G{num_deals+1},D2:D{num_deals+1})/SUM(D2:D{num_deals+1})")
    # Monthly Used Vehicle Gross at summary+5
    for i in range(12):
        ws.cell(row=summary + 5 + i, column=1, value=f"Used Gross {MONTHS[i]}")
        ws.cell(row=summary + 5 + i, column=2,
                value=f"=SUMIFS(I2:I{num_deals+1},J2:J{num_deals+1},\"{MONTHS[i]}\")")


def build_fi_detail(wb: Workbook) -> None:
    ws = wb.create_sheet("F&I Detail")
    headers = ["Deal#", "Product Type", "Product Cost", "Product Revenue", "Reserve", "Month"]
    for i, h in enumerate(headers, 1):
        _header(ws, 1, i, h)
    products = ["Warranty", "GAP", "Tire & Wheel", "Maintenance"]
    for r in range(2, 52):
        deal_id = 1000 + random.randint(2, 42)
        prod = random.choice(products)
        cost = random.randint(200, 900)
        rev = cost + random.randint(300, 1400)
        reserve = random.randint(50, 400)
        ws.cell(row=r, column=1, value=deal_id)
        ws.cell(row=r, column=2, value=prod)
        ws.cell(row=r, column=3, value=cost)
        ws.cell(row=r, column=4, value=rev)
        ws.cell(row=r, column=5, value=reserve)
        ws.cell(row=r, column=6, value=MONTHS[(r - 2) % 12])

    ws.cell(row=55, column=1, value="Total F&I Revenue")
    ws.cell(row=55, column=2, value="=SUM(D2:D51)")
    ws.cell(row=56, column=1, value="Deal Count")
    ws.cell(row=56, column=2, value="=COUNTA(A2:A51)")
    ws.cell(row=57, column=1, value="PVR")
    ws.cell(row=57, column=2, value="=B55/B56")

    for i in range(12):
        ws.cell(row=60 + i, column=1, value=f"F&I Gross {MONTHS[i]}")
        ws.cell(row=60 + i, column=2, value=f"=SUMIFS(D2:D51,F2:F51,\"{MONTHS[i]}\")-SUMIFS(C2:C51,F2:F51,\"{MONTHS[i]}\")")


def build_service_parts(wb: Workbook) -> None:
    ws = wb.create_sheet("Service & Parts")
    headers = ["RO#", "Date", "Type", "Labor Revenue", "Parts Revenue", "Labor Cost", "Parts Cost", "Month"]
    for i, h in enumerate(headers, 1):
        _header(ws, 1, i, h)
    types = ["Customer Pay", "Warranty", "Internal"]
    for r in range(2, 52):
        month = MONTHS[(r - 2) % 12]
        d = date(2026, ((r - 2) % 12) + 1, 5 + (r % 20))
        lr = random.randint(80, 900); pr = random.randint(40, 700)
        lc = round(lr * random.uniform(0.35, 0.55), 2); pc = round(pr * random.uniform(0.55, 0.75), 2)
        ws.cell(row=r, column=1, value=5000 + r)
        ws.cell(row=r, column=2, value=d)
        ws.cell(row=r, column=3, value=random.choice(types))
        ws.cell(row=r, column=4, value=lr)
        ws.cell(row=r, column=5, value=pr)
        ws.cell(row=r, column=6, value=lc)
        ws.cell(row=r, column=7, value=pc)
        ws.cell(row=r, column=8, value=month)

    ws.cell(row=55, column=1, value="Total Labor Gross")
    ws.cell(row=55, column=2, value="=SUM(D2:D51)-SUM(F2:F51)")
    ws.cell(row=56, column=1, value="Total Parts Gross")
    ws.cell(row=56, column=2, value="=SUM(E2:E51)-SUM(G2:G51)")
    ws.cell(row=57, column=1, value="Total Service & Parts Gross")
    ws.cell(row=57, column=2, value="=B55+B56")

    # Monthly S&P gross
    for i in range(12):
        ws.cell(row=60 + i, column=1, value=f"S&P Gross {MONTHS[i]}")
        ws.cell(row=60 + i, column=2,
                value=(f"=(SUMIFS(D2:D51,H2:H51,\"{MONTHS[i]}\")-SUMIFS(F2:F51,H2:H51,\"{MONTHS[i]}\"))"
                       f"+(SUMIFS(E2:E51,H2:H51,\"{MONTHS[i]}\")-SUMIFS(G2:G51,H2:H51,\"{MONTHS[i]}\"))"))

    # Service Absorption Rate — circular reference back to P&L Summary overhead
    ws.cell(row=75, column=1, value="Service Absorption Rate")
    ws.cell(row=75, column=2, value="=B57/'P&L Summary'!B30")


def build_pnl_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("P&L Summary", 0)  # put first
    _write_monthly_header(ws, 1)
    # Revenue line items
    labels = ["New Vehicle Sales", "Used Vehicle Sales", "F&I Income",
              "Service Revenue", "Parts Revenue"]
    for r, label in enumerate(labels, start=2):
        ws.cell(row=r, column=1, value=label)
        for m_idx in range(12):
            ws.cell(row=r, column=m_idx + 2, value=random.randint(30000, 200000))
    # Row 10: Total Revenue
    ws.cell(row=10, column=1, value="Total Revenue")
    for m_idx in range(12):
        col = get_column_letter(m_idx + 2)
        ws.cell(row=10, column=m_idx + 2, value=f"=SUM({col}2:{col}6)")

    # Row 12-17: department gross (pulled from other sheets)
    ws.cell(row=12, column=1, value="New Vehicle Gross")
    for m_idx in range(12):
        # row summary_row + 20 = 45 + 20 = 65 on New Vehicle
        ws.cell(row=12, column=m_idx + 2,
                value=f"='New Vehicle'!B{65 + m_idx}")
    ws.cell(row=13, column=1, value="Used Vehicle Gross")
    for m_idx in range(12):
        ws.cell(row=13, column=m_idx + 2,
                value=f"='Used Vehicle'!B{45 + m_idx}")
    ws.cell(row=14, column=1, value="F&I Gross")
    for m_idx in range(12):
        ws.cell(row=14, column=m_idx + 2,
                value=f"='F&I Detail'!B{60 + m_idx}")
    ws.cell(row=15, column=1, value="Service & Parts Gross")
    for m_idx in range(12):
        ws.cell(row=15, column=m_idx + 2,
                value=f"='Service & Parts'!B{60 + m_idx}")
    # Row 18: Total Gross Profit
    ws.cell(row=18, column=1, value="Total Gross Profit")
    for m_idx in range(12):
        col = get_column_letter(m_idx + 2)
        ws.cell(row=18, column=m_idx + 2,
                value=f"=SUM({col}12:{col}15)")

    # Row 22-28: operating expenses
    exp_labels = ["Salaries", "Floor Plan Interest", "Advertising", "Rent", "Utilities", "Other"]
    for r, label in enumerate(exp_labels, start=22):
        ws.cell(row=r, column=1, value=label)
        for m_idx in range(12):
            ws.cell(row=r, column=m_idx + 2, value=random.randint(5000, 30000))
    ws.cell(row=29, column=1, value="Total Operating Expenses")
    for m_idx in range(12):
        col = get_column_letter(m_idx + 2)
        ws.cell(row=29, column=m_idx + 2, value=f"=SUM({col}22:{col}27)")

    # Row 30: Overhead Base (used in absorption — CIRCULAR source)
    ws.cell(row=30, column=1, value="Total Overhead")
    for m_idx in range(12):
        col = get_column_letter(m_idx + 2)
        # Circular: overhead = operating expenses + a portion of service absorption target
        ws.cell(row=30, column=m_idx + 2,
                value=f"={col}29+OverheadBase*'Service & Parts'!B75")

    # Row 32: Adjusted EBITDA
    ws.cell(row=32, column=1, value="Adjusted EBITDA")
    for m_idx in range(12):
        col = get_column_letter(m_idx + 2)
        ws.cell(row=32, column=m_idx + 2,
                value=f"={col}18-{col}29+OwnerCompensationAddback+OneTimeLegalCosts")

    # Row 35: Net Income after Tax
    ws.cell(row=35, column=1, value="Net Income")
    for m_idx in range(12):
        col = get_column_letter(m_idx + 2)
        ws.cell(row=35, column=m_idx + 2, value=f"={col}32*(1-TaxRate)")


def build_workbook(out_path: str) -> str:
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)
    build_assumptions(wb)
    build_new_vehicle(wb)
    build_used_vehicle(wb)
    build_fi_detail(wb)
    build_service_parts(wb)
    build_pnl_summary(wb)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "data/dealership_financial_model.xlsx"
    p = build_workbook(out)
    print(f"Wrote {p}")
