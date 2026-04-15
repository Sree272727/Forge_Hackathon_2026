"""Generate the Energy portfolio revenue model test workbook.

Highlights:
 - 7 sheets
 - 20+ cross-sheet refs, 10+ named ranges
 - multiplicative system loss formula
 - VLOOKUP-based price curve matching
 - Conditional PPA vs Merchant revenue logic
 - Per-site assumption overrides (Alpha, Gamma)
 - INDIRECT formula for dynamic sheet reference
 - 300+ formulas
"""
from __future__ import annotations

import random
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

random.seed(7)

SITES = ["Alpha", "Beta", "Gamma", "Delta", "Epsilon", "Zeta", "Eta", "Theta", "Iota", "Kappa"]
CONTRACT_TYPES = ["PPA", "Merchant", "Hybrid", "PPA", "PPA", "Merchant", "Hybrid", "PPA", "Merchant", "PPA"]
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _hdr(ws, r, c, v, bold=True):
    ws.cell(row=r, column=c, value=v)
    ws.cell(row=r, column=c).font = Font(bold=bold)


def build_assumptions(wb: Workbook):
    ws = wb.create_sheet("Assumptions")
    headers = ["Assumption", "Value", "Effective Date", "Notes"]
    for i, h in enumerate(headers, 1):
        _hdr(ws, 1, i, h)
    rows = [
        ("AnnualDegradationRate", 0.005, date(2025, 1, 1), "Default annual panel degradation"),
        ("DiscountRate", 0.08, date(2023, 3, 1), "STALE — project WACC"),
        ("ITCRate", 0.30, date(2024, 6, 1), "Investment tax credit rate"),
        ("OM_EscalationRate", 0.025, date(2025, 1, 1), "O&M cost escalation"),
        ("LandLeaseEscalation", 0.02, date(2025, 1, 1), "Land lease escalation"),
        ("InsuranceEscalation", 0.03, date(2023, 1, 1), "STALE — insurance escalation"),
        ("TaxEquitySplit", 0.995, date(2024, 6, 1), "Pre-flip tax equity split"),
        ("PostFlipSplit", 0.05, date(2024, 6, 1), "Post-flip tax equity split"),
        ("TaxEquityFlipDate", date(2028, 6, 1), date(2024, 6, 1), "Flip date"),
        ("DefaultPriceCurveSheet", "Price Curves", date(2025, 1, 1), "Sheet used by INDIRECT"),
        ("PortfolioWideDegradation", 0.005, date(2025, 1, 1), "Also default"),
    ]
    for r, (n, v, dt, note) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=v)
        ws.cell(row=r, column=3, value=dt)
        ws.cell(row=r, column=4, value=note)

    # Per-site overrides (rows 15-16)
    ws.cell(row=15, column=1, value="Alpha_DegradationOverride")
    ws.cell(row=15, column=2, value=0.007)
    ws.cell(row=15, column=3, value=date(2025, 2, 1))
    ws.cell(row=15, column=4, value="Site Alpha override")
    ws.cell(row=16, column=1, value="Gamma_EscalationCap")
    ws.cell(row=16, column=2, value=0.02)
    ws.cell(row=16, column=3, value=date(2025, 2, 1))
    ws.cell(row=16, column=4, value="Site Gamma escalation cap")

    names = [
        ("AnnualDegradationRate", "Assumptions!$B$2"),
        ("DiscountRate", "Assumptions!$B$3"),
        ("ITCRate", "Assumptions!$B$4"),
        ("OM_EscalationRate", "Assumptions!$B$5"),
        ("LandLeaseEscalation", "Assumptions!$B$6"),
        ("InsuranceEscalation", "Assumptions!$B$7"),
        ("TaxEquitySplit", "Assumptions!$B$8"),
        ("TaxEquityFlipDate", "Assumptions!$B$10"),
        ("DefaultPriceCurveSheet", "Assumptions!$B$11"),
        ("Alpha_DegradationOverride", "Assumptions!$B$15"),
        ("Gamma_EscalationCap", "Assumptions!$B$16"),
    ]
    for n, target in names:
        wb.defined_names[n] = DefinedName(name=n, attr_text=target)


def build_generation(wb: Workbook):
    ws = wb.create_sheet("Generation")
    _hdr(ws, 1, 1, "Site")
    _hdr(ws, 1, 2, "Inverter")
    _hdr(ws, 1, 3, "Month")
    _hdr(ws, 1, 4, "kWh Generated")
    _hdr(ws, 1, 5, "Availability")
    _hdr(ws, 1, 6, "Curtailment kWh")
    _hdr(ws, 1, 7, "Net Generation")
    r = 2
    for site_idx, site in enumerate(SITES):
        for inv in range(1, 3):
            for m_idx, m in enumerate(MONTHS):
                ws.cell(row=r, column=1, value=site)
                ws.cell(row=r, column=2, value=f"INV-{site[:2].upper()}-{inv}")
                ws.cell(row=r, column=3, value=m)
                ws.cell(row=r, column=4, value=random.randint(50_000, 200_000))
                ws.cell(row=r, column=5, value=round(random.uniform(0.95, 0.995), 4))
                ws.cell(row=r, column=6, value=random.randint(0, 3000))
                ws.cell(row=r, column=7, value=f"=D{r}-F{r}")
                r += 1
    # Site totals at bottom
    last_row = r - 1
    ws.cell(row=r, column=1, value="TOTAL BY SITE")
    for i, site in enumerate(SITES):
        ws.cell(row=r + 1 + i, column=1, value=site)
        ws.cell(row=r + 1 + i, column=4,
                value=f"=SUMIFS(G2:G{last_row},A2:A{last_row},\"{site}\")")


def build_losses(wb: Workbook):
    ws = wb.create_sheet("System Losses")
    headers = ["Site", "Soiling", "Shading", "Wiring", "Transformer", "Availability", "Degradation", "Total Loss"]
    for i, h in enumerate(headers, 1):
        _hdr(ws, 1, i, h)
    for i, site in enumerate(SITES):
        r = 2 + i
        ws.cell(row=r, column=1, value=site)
        ws.cell(row=r, column=2, value=round(random.uniform(0.01, 0.03), 4))
        ws.cell(row=r, column=3, value=round(random.uniform(0.005, 0.02), 4))
        ws.cell(row=r, column=4, value=round(random.uniform(0.005, 0.015), 4))
        ws.cell(row=r, column=5, value=round(random.uniform(0.01, 0.02), 4))
        ws.cell(row=r, column=6, value=round(random.uniform(0.005, 0.02), 4))
        # Degradation uses per-site override for Alpha, default for others
        if site == "Alpha":
            ws.cell(row=r, column=7, value="=Alpha_DegradationOverride")
        else:
            ws.cell(row=r, column=7, value="=AnnualDegradationRate")
        # Multiplicative total: 1 - PRODUCT(1-loss_i)
        ws.cell(row=r, column=8, value=f"=1-PRODUCT(1-B{r},1-C{r},1-D{r},1-E{r},1-F{r},1-G{r})")


def build_weather(wb: Workbook):
    ws = wb.create_sheet("Weather")
    _hdr(ws, 1, 1, "Site")
    for i, m in enumerate(MONTHS, start=2):
        _hdr(ws, 1, i, f"TMY_{m}")
    for i, m in enumerate(MONTHS, start=14):
        _hdr(ws, 1, i, f"Actual_{m}")
    _hdr(ws, 1, 26, "WeatherAdj")
    for i, site in enumerate(SITES):
        r = 2 + i
        ws.cell(row=r, column=1, value=site)
        for j in range(12):
            ws.cell(row=r, column=2 + j, value=round(random.uniform(120, 200), 1))
            ws.cell(row=r, column=14 + j, value=round(random.uniform(110, 210), 1))
        ws.cell(row=r, column=26, value=f"=SUM(N{r}:Y{r})/SUM(B{r}:M{r})")


def build_price_curves(wb: Workbook):
    ws = wb.create_sheet("Price Curves")
    _hdr(ws, 1, 1, "Month")
    _hdr(ws, 1, 2, "Merchant Price")
    _hdr(ws, 1, 3, "Hybrid Floor")
    for i, m in enumerate(MONTHS):
        r = 2 + i
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=round(random.uniform(45, 85), 2))
        ws.cell(row=r, column=3, value=round(random.uniform(35, 55), 2))

    # Dynamic reference demo via INDIRECT
    ws.cell(row=20, column=1, value="Dynamic Jan Merchant (via INDIRECT)")
    ws.cell(row=20, column=2, value='=INDIRECT(DefaultPriceCurveSheet&"!B2")')


def build_costs(wb: Workbook):
    ws = wb.create_sheet("O&M Costs")
    headers = ["Site", "O&M Annual", "Insurance", "Land Lease", "Property Tax", "Asset Mgmt Fee", "Total Opex"]
    for i, h in enumerate(headers, 1):
        _hdr(ws, 1, i, h)
    for i, site in enumerate(SITES):
        r = 2 + i
        ws.cell(row=r, column=1, value=site)
        base_om = random.randint(80_000, 250_000)
        ws.cell(row=r, column=2, value=base_om)
        ws.cell(row=r, column=3, value=random.randint(10_000, 40_000))
        ws.cell(row=r, column=4, value=random.randint(20_000, 60_000))
        ws.cell(row=r, column=5, value=random.randint(15_000, 45_000))
        ws.cell(row=r, column=6, value=random.randint(8_000, 20_000))
        ws.cell(row=r, column=7, value=f"=SUM(B{r}:F{r})")
    # Escalated next year (applies rates)
    ws.cell(row=13, column=1, value="NEXT YEAR OPEX")
    for i, site in enumerate(SITES):
        r = 14 + i
        src = 2 + i
        ws.cell(row=r, column=1, value=site)
        if site == "Gamma":
            ws.cell(row=r, column=7, value=f"=G{src}*(1+Gamma_EscalationCap)")
        else:
            ws.cell(row=r, column=7, value=f"=G{src}*(1+OM_EscalationRate)")


def build_portfolio_summary(wb: Workbook):
    ws = wb.create_sheet("Portfolio Summary", 0)
    headers = ["Site", "Capacity (MW)", "Expected Yield (MWh)", "Contract Type",
               "Contract Rate", "Expected Revenue", "System Loss", "Performance Ratio",
               "Net Operating Income"]
    for i, h in enumerate(headers, 1):
        _hdr(ws, 1, i, h)
    for i, site in enumerate(SITES):
        r = 2 + i
        ct = CONTRACT_TYPES[i]
        cap = round(random.uniform(5, 50), 1)
        yld = round(cap * random.uniform(1600, 2100), 0)
        rate = round(random.uniform(55, 90), 2)
        ws.cell(row=r, column=1, value=site)
        ws.cell(row=r, column=2, value=cap)
        ws.cell(row=r, column=3, value=yld)
        ws.cell(row=r, column=4, value=ct)
        ws.cell(row=r, column=5, value=rate)
        # Expected Revenue — conditional on contract type
        # For PPA: Yield * Rate
        # For Merchant: Yield * VLOOKUP first month merchant price
        # For Hybrid: Yield * MAX(merchant, floor)
        ws.cell(row=r, column=6,
                value=(f"=IF(D{r}=\"PPA\",C{r}*E{r},"
                       f"IF(D{r}=\"Merchant\",C{r}*VLOOKUP(\"Jan\",'Price Curves'!A:B,2,FALSE),"
                       f"C{r}*MAX(VLOOKUP(\"Jan\",'Price Curves'!A:B,2,FALSE),VLOOKUP(\"Jan\",'Price Curves'!A:C,3,FALSE))))"))
        # System Loss lookup
        ws.cell(row=r, column=7, value=f"=VLOOKUP(A{r},'System Losses'!A:H,8,FALSE)")
        # Performance Ratio = (1 - System Loss) * Weather Adj
        ws.cell(row=r, column=8, value=f"=(1-G{r})*VLOOKUP(A{r},Weather!A:Z,26,FALSE)")
        # NOI = Expected Revenue - Total Opex
        ws.cell(row=r, column=9, value=f"=F{r}-VLOOKUP(A{r},'O&M Costs'!A:G,7,FALSE)")

    # Portfolio totals
    ws.cell(row=13, column=1, value="PORTFOLIO TOTAL")
    ws.cell(row=13, column=2, value="=SUM(B2:B11)")
    ws.cell(row=13, column=3, value="=SUM(C2:C11)")
    ws.cell(row=13, column=6, value="=SUM(F2:F11)")
    ws.cell(row=13, column=9, value="=SUM(I2:I11)")
    ws.cell(row=14, column=1, value="Weighted Avg PR")
    ws.cell(row=14, column=8, value="=SUMPRODUCT(B2:B11,H2:H11)/SUM(B2:B11)")


def build_workbook(out_path: str) -> str:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    # Order: Assumptions first so names resolve to existing cells, but Portfolio Summary shown first visually
    build_assumptions(wb)
    build_generation(wb)
    build_losses(wb)
    build_weather(wb)
    build_price_curves(wb)
    build_costs(wb)
    build_portfolio_summary(wb)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "data/energy_portfolio_model.xlsx"
    p = build_workbook(out)
    print(f"Wrote {p}")
